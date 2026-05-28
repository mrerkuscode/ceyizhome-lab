from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from template_writer import PRODUCTION_COLUMNS
from text_utils import clean_customer_name, normalize_cell


LEGACY_COLUMN_ALIASES = {
    "order_no": {
        "siparis_numarasi",
        "siparis_no",
        "siparis",
        "order_no",
        "order_number",
    },
    "buyer_name": {
        "alici_ismi",
        "alici_ismi",
        "alici",
        "alici_adi",
        "musteri",
        "musteri_adi",
        "buyer_name",
    },
    "gold": {
        "gold",
        "altin",
        "gold_yazi",
        "altin_yazi",
    },
    "silver": {
        "gumus",
        "gümüs",
        "gümüş",
        "silver",
        "gumus_yazi",
        "silver_yazi",
    },
    "product_name": {
        "urun",
        "urun_adi",
        "urun_ismi",
        "product",
        "product_name",
    },
    "note": {
        "not",
        "notlar",
        "aciklama",
        "note",
        "production_note",
    },
    "quantity": {
        "adet",
        "quantity",
        "qty",
        "miktar",
    },
}

NO_NAME_PATTERNS = ("isim yok", "isimsiz", "kişiselleştirme yok", "kisisellestirme yok", "yazı yok", "yazi yok")
NOTE_KEYWORDS = (
    "not",
    "acil",
    "dikkat",
    "renk",
    "font",
    "logo",
    "tekrar",
    "kontrol",
    "paket",
    "teslim",
)


@dataclass(frozen=True)
class LegacyConversionResult:
    clean_excel_path: Path
    normalized_csv_path: Path
    warnings_csv_path: Path
    converted_rows: int
    warning_rows: int


def convert_legacy_excel(input_path: Path, output_dir: Path) -> LegacyConversionResult:
    if not input_path.exists():
        raise FileNotFoundError(f"Legacy Excel file not found: {input_path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    raw_dataframe = pd.read_excel(input_path, dtype=object, engine="openpyxl")
    column_map = _build_column_map(raw_dataframe.columns)

    clean_rows: list[dict[str, Any]] = []
    warning_rows: list[dict[str, str]] = []
    current_product_group = ""

    for index, row in raw_dataframe.iterrows():
        excel_row_number = int(index) + 2
        legacy = _extract_legacy_row(row, column_map)

        if _is_blank_legacy_row(legacy):
            continue

        if _is_product_group_header(legacy):
            current_product_group = legacy["product_name"] or legacy["note"]
            warning_rows.append(
                _warning(
                    excel_row_number,
                    "",
                    "product_group_header",
                    f"Ürün grup başlığı algılandı: {current_product_group}",
                )
            )
            continue

        clean_row, row_warnings = _convert_order_row(legacy, excel_row_number, current_product_group)
        clean_rows.append(clean_row)
        warning_rows.extend(row_warnings)

    clean_excel_path = output_dir / "cyzella_clean_orders.xlsx"
    normalized_csv_path = output_dir / "normalized_orders.csv"
    warnings_csv_path = output_dir / "legacy_import_warnings.csv"

    _write_clean_excel(clean_excel_path, clean_rows)
    _write_normalized_csv(normalized_csv_path, clean_rows)
    _write_warnings_csv(warnings_csv_path, warning_rows)

    return LegacyConversionResult(
        clean_excel_path=clean_excel_path,
        normalized_csv_path=normalized_csv_path,
        warnings_csv_path=warnings_csv_path,
        converted_rows=len(clean_rows),
        warning_rows=len(warning_rows),
    )


def _build_column_map(columns: list[object]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    normalized_to_original = {_normalize_legacy_column(column): str(column) for column in columns}

    for target, aliases in LEGACY_COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized_to_original:
                mapping[target] = normalized_to_original[alias]
                break
    return mapping


def _extract_legacy_row(row: pd.Series, column_map: dict[str, str]) -> dict[str, str]:
    return {
        "order_no": _legacy_cell(row, column_map, "order_no"),
        "buyer_name": clean_customer_name(_legacy_cell(row, column_map, "buyer_name")),
        "gold": clean_customer_name(_legacy_cell(row, column_map, "gold")),
        "silver": clean_customer_name(_legacy_cell(row, column_map, "silver")),
        "product_name": clean_customer_name(_legacy_cell(row, column_map, "product_name")),
        "note": clean_customer_name(_legacy_cell(row, column_map, "note")),
        "quantity": _legacy_cell(row, column_map, "quantity"),
    }


def _legacy_cell(row: pd.Series, column_map: dict[str, str], key: str) -> str:
    source_column = column_map.get(key)
    if not source_column:
        return ""
    return normalize_cell(row.get(source_column))


def _is_blank_legacy_row(legacy: dict[str, str]) -> bool:
    return not any(value.strip() for value in legacy.values())


def _is_product_group_header(legacy: dict[str, str]) -> bool:
    has_product_text = bool(legacy["product_name"] or legacy["note"])
    has_order_data = bool(legacy["order_no"] or legacy["buyer_name"] or legacy["gold"] or legacy["silver"])
    return has_product_text and not has_order_data


def _convert_order_row(
    legacy: dict[str, str],
    row_number: int,
    current_product_group: str,
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    warnings: list[dict[str, str]] = []
    order_no = _normalize_order_no(legacy["order_no"]) or f"LEGACY-ROW-{row_number}"
    buyer_name = legacy["buyer_name"]
    product_name = legacy["product_name"] or current_product_group
    production_note_parts = [legacy["note"]]
    extra_chocolate_qty = _extract_extra_qty(product_name, "çikolata", "cikolata")
    extra_madlen_qty = _extract_extra_qty(product_name, "madlen")

    label_variant, label_text, variant_warnings = _extract_label_text(legacy)
    warnings.extend(_warning(row_number, order_no, "label_text", message) for message in variant_warnings)

    no_personalization = _is_no_personalization_text(legacy["gold"]) or _is_no_personalization_text(legacy["silver"])
    if no_personalization:
        label_text = ""
        label_variant = "NONE"

    needs_review_reasons: list[str] = []
    if order_no.startswith("LEGACY-ROW"):
        needs_review_reasons.append("sipariş numarası eksik")
    if not buyer_name:
        needs_review_reasons.append("alıcı ismi eksik")
    if not product_name:
        needs_review_reasons.append("ürün adı eksik")

    process_type = "PRINT" if label_text else "NONE"
    personalization_type = "LABEL" if label_text else "NO_PERSONALIZATION"
    if not no_personalization:
        needs_review_reasons.extend(["model_no belirsiz", "template_no belirsiz", "process_type eski Excel'den kesin doğrulanamadı"])

    if _looks_like_note(legacy["gold"]):
        production_note_parts.append(f"Gold alanında not olabilir: {legacy['gold']}")
        needs_review_reasons.append("gold alanı not gibi görünüyor")
    if _looks_like_note(legacy["silver"]):
        production_note_parts.append(f"Gümüş alanında not olabilir: {legacy['silver']}")
        needs_review_reasons.append("gümüş alanı not gibi görünüyor")

    if extra_chocolate_qty:
        warnings.append(_warning(row_number, order_no, "extra_chocolate_qty", f"Ekstra çikolata algılandı: {extra_chocolate_qty}"))
    if extra_madlen_qty:
        warnings.append(_warning(row_number, order_no, "extra_madlen_qty", f"Ekstra madlen algılandı: {extra_madlen_qty}"))

    if needs_review_reasons:
        warnings.append(_warning(row_number, order_no, "needs_review", "; ".join(sorted(set(needs_review_reasons)))))

    clean_row = {
        "order_no": order_no,
        "buyer_name": buyer_name,
        "product_name": product_name,
        "model_no": "",
        "template_no": "",
        "process_type": process_type,
        "personalization_type": personalization_type,
        "label_variant": label_variant,
        "label_text": label_text,
        "laser_text": "",
        "quantity": _parse_quantity(legacy["quantity"]),
        "material_type": "",
        "material_thickness_mm": "",
        "extra_chocolate_qty": extra_chocolate_qty,
        "extra_madlen_qty": extra_madlen_qty,
        "production_note": " | ".join(part for part in production_note_parts if part),
        "needs_review": "true" if needs_review_reasons else "",
        "status": "NEEDS_REVIEW" if needs_review_reasons else "NEW",
    }
    return clean_row, warnings


def _extract_label_text(legacy: dict[str, str]) -> tuple[str, str, list[str]]:
    gold = "" if _is_no_personalization_text(legacy["gold"]) else legacy["gold"]
    silver = "" if _is_no_personalization_text(legacy["silver"]) else legacy["silver"]
    warnings: list[str] = []

    if gold and silver:
        warnings.append("Hem gold hem gümüş alanı dolu; kullanıcı kontrol etmeli.")
        return "CUSTOM", f"GOLD: {gold} | SILVER: {silver}", warnings
    if gold:
        return "GOLD", gold, warnings
    if silver:
        return "SILVER", silver, warnings
    return "NONE", "", warnings


def _normalize_order_no(value: str) -> str:
    text = normalize_cell(value)
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\s+", "", text)


def _parse_quantity(value: str) -> int:
    if not value:
        return 1
    try:
        parsed = int(float(str(value).replace(",", ".")))
        return parsed if parsed > 0 else 1
    except ValueError:
        return 1


def _extract_extra_qty(text: str, *keywords: str) -> int:
    normalized = _normalize_text(text)
    if not any(keyword in normalized for keyword in keywords):
        return 0
    patterns = [
        r"(\d+)\s*(:adet|ad|x)\s*(:" + "|".join(re.escape(keyword) for keyword in keywords) + r")",
        r"(:" + "|".join(re.escape(keyword) for keyword in keywords) + r")\s*(\d+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, normalized)
        if match:
            return int(match.group(1))
    return 1


def _is_no_personalization_text(value: str) -> bool:
    normalized = _normalize_text(value)
    return any(pattern in normalized for pattern in NO_NAME_PATTERNS)


def _looks_like_note(value: str) -> bool:
    normalized = _normalize_text(value)
    if not normalized:
        return False
    if _is_no_personalization_text(value):
        return False
    return any(keyword in normalized for keyword in NOTE_KEYWORDS) or len(normalized.split()) > 5


def _normalize_legacy_column(value: object) -> str:
    return _normalize_text(str(value)).strip("_")


def _normalize_text(value: str) -> str:
    text = value.strip().lower()
    replacements = str.maketrans(
        {
            "ç": "c",
            "ğ": "g",
            "ı": "i",
            "ö": "o",
            "ş": "s",
            "ü": "u",
        }
    )
    text = text.translate(replacements)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def _warning(row_number: int, order_no: str, field: str, message: str) -> dict[str, str]:
    return {
        "legacy_row_number": str(row_number),
        "order_no": order_no,
        "field": field,
        "message": message,
    }


def _write_clean_excel(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clean Orders"
    sheet.append(PRODUCTION_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in PRODUCTION_COLUMNS])

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def _write_normalized_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=PRODUCTION_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in PRODUCTION_COLUMNS})


def _write_warnings_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["legacy_row_number", "order_no", "field", "message"],
        )
        writer.writeheader()
        writer.writerows(rows)
