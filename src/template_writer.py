from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PRODUCTION_COLUMNS = [
    "order_no",
    "buyer_name",
    "product_name",
    "model_no",
    "template_no",
    "process_type",
    "personalization_type",
    "label_variant",
    "label_text",
    "laser_text",
    "quantity",
    "material_type",
    "material_thickness_mm",
    "extra_chocolate_qty",
    "extra_madlen_qty",
    "production_note",
    "needs_review",
    "status",
]

TURKISH_PRODUCTION_HEADERS = {
    "order_no": "Sipariş No",
    "buyer_name": "Alıcı Adı",
    "product_name": "Ürün Adı",
    "model_no": "Model No",
    "template_no": "Şablon No",
    "process_type": "İşlem Tipi",
    "personalization_type": "Kişiselleştirme Türü",
    "label_variant": "Etiket Varyantı",
    "label_text": "Etiket Yazısı",
    "laser_text": "Lazer Yazısı",
    "quantity": "Adet",
    "material_type": "Malzeme Türü",
    "material_thickness_mm": "Malzeme Kalınlığı mm",
    "extra_chocolate_qty": "Ekstra Çikolata Adedi",
    "extra_madlen_qty": "Ekstra Madlen Adedi",
    "production_note": "Üretim Notu",
    "needs_review": "Kontrol Gerekli",
    "status": "Durum",
}

PROCESS_TYPES = ["PRINT", "LASER_ENGRAVE", "LASER_CUT", "BOTH", "NONE"]
PERSONALIZATION_TYPES = ["LABEL", "NAME", "LABEL_AND_NAME", "NO_PERSONALIZATION"]
LABEL_VARIANTS = ["GOLD", "SILVER", "WHITE", "RED", "CUSTOM", "NONE"]
STATUSES = ["NEW", "READY", "NEEDS_REVIEW", "COMPLETED", "CANCELLED"]

SAMPLE_ROWS = [
    [
        "CYZ-1001",
        "Ayşe",
        "Gold Çikolata Etiketi",
        "01",
        "A",
        "PRINT",
        "LABEL",
        "GOLD",
        "Ayşe",
        "",
        24,
        "",
        "",
        0,
        0,
        "Gold baskı etiketi hazırlanacak.",
        "",
        "NEW",
    ],
    [
        "CYZ-1002",
        "Gülşah",
        "Silver Çikolata Etiketi",
        "01",
        "A",
        "PRINT",
        "LABEL",
        "SILVER",
        "Gülşah",
        "",
        18,
        "",
        "",
        0,
        0,
        "Silver baskı etiketi hazırlanacak.",
        "",
        "NEW",
    ],
    [
        "CYZ-1003",
        "Mücahit",
        "Bağlantılı El Yazısı İsim",
        "07",
        "NAME",
        "LASER_CUT",
        "NAME",
        "NONE",
        "",
        "Mücahit",
        1,
        "Acrylic",
        3,
        0,
        0,
        "Lazer kesim bağlı el yazısı isim.",
        "",
        "NEW",
    ],
    [
        "CYZ-1004",
        "İrem",
        "Etiket ve Lazer Kazıma Seti",
        "12",
        "B",
        "BOTH",
        "LABEL_AND_NAME",
        "WHITE",
        "İrem",
        "İrem",
        12,
        "Wood",
        4,
        0,
        0,
        "Hem etiket hem lazer kazıma hazırlanacak.",
        "",
        "NEW",
    ],
    [
        "CYZ-1005",
        "Çağla",
        "Hazır Ürün",
        "20",
        "STD",
        "NONE",
        "NO_PERSONALIZATION",
        "NONE",
        "",
        "",
        6,
        "",
        "",
        0,
        0,
        "Kişiselleştirme yok.",
        "",
        "NEW",
    ],
    [
        "CYZ-1006",
        "Ömer",
        "Ekstra Çikolatalı Set",
        "03",
        "A",
        "PRINT",
        "LABEL",
        "RED",
        "Ömer",
        "",
        10,
        "",
        "",
        8,
        0,
        "Ekstra çikolata eklenecek.",
        "",
        "NEW",
    ],
    [
        "CYZ-1007",
        "Şükran",
        "Ekstra Madlenli Set",
        "04",
        "A",
        "PRINT",
        "LABEL",
        "CUSTOM",
        "Şükran",
        "",
        10,
        "",
        "",
        0,
        12,
        "Ekstra madlen eklenecek.",
        "",
        "NEW",
    ],
    [
        "CYZ-1008",
        "Ayşe",
        "Lazer Kazıma İsim",
        "08",
        "ENG",
        "LASER_ENGRAVE",
        "NAME",
        "NONE",
        "",
        "Ayşe & Mehmet",
        2,
        "Acrylic",
        3,
        0,
        0,
        "Lazer kazıma örneği.",
        "",
        "NEW",
    ],
    [
        "CYZ-1009",
        "Gülşah",
        "Eksik Şablonlu Gold Etiket",
        "99",
        "MISSING",
        "PRINT",
        "LABEL",
        "GOLD",
        "Gülşah",
        "",
        5,
        "",
        "",
        0,
        0,
        "Bu satır missing template raporu için bilerek eklendi.",
        "",
        "NEW",
    ],
    [
        "CYZ-1010",
        "Mücahit",
        "Eksik Lazer Metni",
        "09",
        "NAME",
        "LASER_CUT",
        "NAME",
        "NONE",
        "",
        "",
        1,
        "Acrylic",
        3,
        0,
        0,
        "Bu satır missing laser_text raporu için bilerek eklendi.",
        "",
        "NEW",
    ],
    [
        "CYZ-1011",
        "İrem",
        "Çok Uzun Lazer Kazıma",
        "10",
        "ENG",
        "LASER_ENGRAVE",
        "NAME",
        "NONE",
        "",
        "İrem Çağla Şükran Ayşe Gülşah Mücahit Ömer Çok Uzun Lazer Yazısı Deneme",
        1,
        "Acrylic",
        3,
        0,
        0,
        "Bu satır long laser_text kontrolü için eklendi.",
        "",
        "NEW",
    ],
]


def create_production_template(output_path: Path) -> Path:
    return _create_workbook(output_path, "Production Orders", [])


def create_demo_orders(output_path: Path) -> Path:
    return _create_workbook(output_path, "Demo Orders", SAMPLE_ROWS)


def _create_workbook(output_path: Path, sheet_title: str, rows: list[list[object]]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.append([TURKISH_PRODUCTION_HEADERS[column] for column in PRODUCTION_COLUMNS])

    for row in rows:
        sheet.append(row)

    _style_header(sheet)
    _set_column_widths(sheet)
    _add_dropdowns(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = _available_output_path(output_path)
        workbook.save(fallback_path)
        return fallback_path


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def _set_column_widths(sheet) -> None:
    widths = {
        "A": 16,
        "B": 18,
        "C": 28,
        "D": 12,
        "E": 14,
        "F": 18,
        "G": 24,
        "H": 16,
        "I": 24,
        "J": 24,
        "K": 10,
        "L": 18,
        "M": 22,
        "N": 20,
        "O": 18,
        "P": 36,
        "Q": 16,
        "R": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _add_dropdowns(sheet) -> None:
    dropdowns = {
        "F2:F500": PROCESS_TYPES,
        "G2:G500": PERSONALIZATION_TYPES,
        "H2:H500": LABEL_VARIANTS,
        "R2:R500": STATUSES,
    }

    for cell_range, values in dropdowns.items():
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(values)}"',
            allow_blank=False,
        )
        validation.error = "Lütfen listeden geçerli bir değer seçin."
        validation.errorTitle = "Geçersiz değer"
        validation.prompt = "Listeden bir değer seçin."
        validation.promptTitle = "İzin verilen değerler"
        sheet.add_data_validation(validation)
        validation.add(cell_range)


def _available_output_path(output_path: Path) -> Path:
    for index in range(1, 100):
        candidate = output_path.with_name(f"{output_path.stem}_yeni_{index}{output_path.suffix}")
        if not candidate.exists():
            return candidate
    raise PermissionError(f"Excel dosyası açık olduğu için şablon kaydedilemedi: {output_path}")
